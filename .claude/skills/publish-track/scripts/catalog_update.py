#!/usr/bin/env python3
"""
catalog_update.py — WoP Media Catalog Manager
================================================
Add, update, and query entries in WoP_Media_Catalog.xlsx.

Designed for the 39-column catalog format (v2) with grouped headers:
  Identity (1-7) | Content (8-11) | Credits (12-16) | Metadata (17-20)
  Source (21-24)  | Tiers (25-29)  | Deployment (30-35) | Pipeline (36-39)

Usage:
  # Add a new music track (auto-generates Catalog ID)
  python catalog_update.py add \
    --title "I Have Tasted the Light" \
    --prefix 06 --chapter 6 --track 1 --version 1 \
    --style "Sacred Americana" \
    --source-platform "Suno AI" \
    --pipeline "Complete"

  # Add a narration track
  python catalog_update.py add \
    --title "Introduction to Plainness" \
    --media-type Narration \
    --prefix NR --chapter 1 --track 1 \
    --narrator "Aaron Powner" \
    --album "Words of Plainness: Chapter Readings"

  # Update an existing entry by Catalog ID
  python catalog_update.py update WOP-0005 \
    --pipeline "Complete" \
    --distrokid "✓" \
    --streaming-live "✓"

  # Update by matching title + style (for batch operations)
  python catalog_update.py update-by-title \
    --title "Introduction to Plainness" \
    --style "Sacred Americana" \
    --web-mp3 "✓" --vbr-v2 Y --website-live "✓"

  # List all entries (summary view)
  python catalog_update.py list

  # List entries filtered by prefix or status
  python catalog_update.py list --prefix 06
  python catalog_update.py list --pipeline "Legacy — Needs Pipeline"

  # Show next available Catalog ID
  python catalog_update.py next-id

  # Export catalog to CSV (for review)
  python catalog_update.py export catalog_export.csv

Part of the /publish pipeline.
Location: .claude/skills/publish-track/scripts/catalog_update.py
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ===========================================================================
# COLUMN MAP — 1-indexed positions matching WoP_Media_Catalog.xlsx (v2)
# ===========================================================================
COL = {
    # Identity & Classification (1-7)
    "catalog_id":       1,
    "isrc":             2,
    "media_type":       3,
    "prefix":           4,
    "chapter":          5,
    "track":            6,
    "version":          7,
    # Content (8-11)
    "title":            8,
    "style":            9,
    "duration":         10,
    "description":      11,
    # Credits (12-16)
    "artist":           12,
    "composer":         13,
    "narrator":         14,
    "copyright":        15,
    "year":             16,
    # Technical Metadata (17-20)
    "album":            17,
    "genre":            18,
    "bpm":              19,
    "key":              20,
    # Source & Production (21-24)
    "source_platform":  21,
    "source_file":      22,
    "mastered":         23,
    "mastering_tool":   24,
    # Archive Tiers (25-29)
    "archive_wav":      25,
    "distro_wav":       26,
    "web_mp3":          27,
    "vbr_v2":           28,
    "album_art":        29,
    # Deployment (30-35)
    "web_filename":     30,
    "website_live":     31,
    "r2_cdn":           32,
    "distrokid":        33,
    "release_date":     34,
    "streaming_live":   35,
    # Pipeline & Status (36-39)
    "pipeline":         36,
    "id3_tagged":       37,
    "last_updated":     38,
    "notes":            39,
}

# Defaults applied to every new entry
DEFAULTS = {
    "media_type":   "Music",
    "artist":       "Words of Plainness",
    "composer":     "Aaron J Powner",
    "copyright":    "© 2026 Aaron J Powner",
    "year":         2026,
    "genre":        "Christian / Sacred",
    "album":        "Words of Plainness: Musical Testimonies",
    "version":      1,
}

# Album defaults per media type
ALBUM_DEFAULTS = {
    "Music":        "Words of Plainness: Musical Testimonies",
    "Narration":    "Words of Plainness: Chapter Readings",
    "Podcast":      "Words of Plainness: Conversations",
    "Spoken Word":  "Words of Plainness: Musical Testimonies",
    "Ambient":      "Words of Plainness: Musical Testimonies",
    "Video":        "Words of Plainness: Musical Testimonies",
    "Other":        "Words of Plainness",
}

# Data row styling
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F7F9FC")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Columns that should be center-aligned
CENTER_COLS = {
    COL["catalog_id"], COL["isrc"], COL["media_type"], COL["prefix"],
    COL["chapter"], COL["track"], COL["version"], COL["duration"],
    COL["year"], COL["bpm"], COL["key"], COL["mastered"],
    COL["archive_wav"], COL["distro_wav"], COL["web_mp3"],
    COL["vbr_v2"], COL["album_art"], COL["website_live"],
    COL["r2_cdn"], COL["distrokid"], COL["release_date"],
    COL["streaming_live"], COL["pipeline"], COL["id3_tagged"],
    COL["last_updated"],
}


# ===========================================================================
# CATALOG FILE LOCATION
# ===========================================================================
def find_catalog() -> Path:
    """Locate WoP_Media_Catalog.xlsx, searching common locations."""
    candidates = [
        Path("WoP_Media_Catalog.xlsx"),
        Path(".claude/skills/publish-track/WoP_Media_Catalog.xlsx"),
        Path("../WoP_Media_Catalog.xlsx"),
    ]
    # Also check CATALOG_PATH env var
    env_path = os.environ.get("WOP_CATALOG_PATH")
    if env_path:
        candidates.insert(0, Path(env_path))

    for p in candidates:
        if p.exists():
            return p.resolve()

    print("Error: WoP_Media_Catalog.xlsx not found.")
    print("Searched:", [str(c) for c in candidates])
    print("Set WOP_CATALOG_PATH env var or run from the project root.")
    sys.exit(1)


# ===========================================================================
# HELPERS
# ===========================================================================
def get_data_start_row() -> int:
    """Row 1 = group headers, Row 2 = column headers, Row 3+ = data."""
    return 3


def next_catalog_id(ws) -> str:
    """Generate the next WOP-NNNN ID based on existing entries."""
    max_id = 0
    for row in range(get_data_start_row(), ws.max_row + 1):
        val = ws.cell(row=row, column=COL["catalog_id"]).value
        if val and isinstance(val, str) and val.startswith("WOP-"):
            try:
                num = int(val.split("-")[1])
                max_id = max(max_id, num)
            except (ValueError, IndexError):
                pass
    return f"WOP-{max_id + 1:04d}"


def find_next_empty_row(ws) -> int:
    """Find the first row where Catalog ID is empty."""
    for row in range(get_data_start_row(), ws.max_row + 2):
        if not ws.cell(row=row, column=COL["catalog_id"]).value:
            return row
    return ws.max_row + 1


def find_row_by_id(ws, catalog_id: str) -> int | None:
    """Find row number for a given Catalog ID."""
    for row in range(get_data_start_row(), ws.max_row + 1):
        val = ws.cell(row=row, column=COL["catalog_id"]).value
        if val and str(val).strip().upper() == catalog_id.strip().upper():
            return row
    return None


def find_row_by_title_style(ws, title: str, style: str = None) -> int | None:
    """Find row by title (and optionally style descriptor)."""
    for row in range(get_data_start_row(), ws.max_row + 1):
        row_title = ws.cell(row=row, column=COL["title"]).value
        if not row_title:
            continue
        if str(row_title).strip().lower() == title.strip().lower():
            if style:
                row_style = ws.cell(row=row, column=COL["style"]).value
                if row_style and str(row_style).strip().lower() == style.strip().lower():
                    return row
            else:
                return row
    return None


def apply_cell_style(ws, row: int, col: int):
    """Apply consistent styling to a data cell."""
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER if col in CENTER_COLS else LEFT
    # Alternating row shading
    if row % 2 == 0:
        cell.fill = ALT_FILL


def set_cell(ws, row: int, col_name: str, value, style=True):
    """Set a cell value by column name and optionally apply styling."""
    col = COL[col_name]
    ws.cell(row=row, column=col).value = value
    if style:
        apply_cell_style(ws, row, col)


def get_cell(ws, row: int, col_name: str):
    """Get a cell value by column name."""
    return ws.cell(row=row, column=COL[col_name]).value


def stamp_updated(ws, row: int):
    """Set the Last Updated field to today."""
    set_cell(ws, row, "last_updated", datetime.now().strftime("%Y-%m-%d"))


# ===========================================================================
# COMMANDS
# ===========================================================================
def cmd_add(args):
    """Add a new entry to the catalog."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path)
    ws = wb["Media Catalog"]

    row = find_next_empty_row(ws)
    cat_id = next_catalog_id(ws)
    media_type = args.media_type or DEFAULTS["media_type"]

    # Check for duplicate title+style
    existing = find_row_by_title_style(ws, args.title, args.style)
    if existing:
        print(f"Warning: Entry already exists at row {existing} "
              f"('{get_cell(ws, existing, 'title')}' / "
              f"'{get_cell(ws, existing, 'style')}')")
        if not args.force:
            print("Use --force to add a duplicate, or use 'update' command instead.")
            return

    # Identity
    set_cell(ws, row, "catalog_id", cat_id)
    if args.isrc:
        set_cell(ws, row, "isrc", args.isrc)
    set_cell(ws, row, "media_type", media_type)
    set_cell(ws, row, "prefix", args.prefix)
    if args.chapter:
        set_cell(ws, row, "chapter", int(args.chapter))
    if args.track:
        set_cell(ws, row, "track", int(args.track))
    set_cell(ws, row, "version", int(args.version or DEFAULTS["version"]))

    # Content
    set_cell(ws, row, "title", args.title)
    if args.style:
        set_cell(ws, row, "style", args.style)
    if args.duration:
        set_cell(ws, row, "duration", args.duration)
    if args.description:
        set_cell(ws, row, "description", args.description)

    # Credits
    set_cell(ws, row, "artist", args.artist or DEFAULTS["artist"])
    set_cell(ws, row, "composer", args.composer or DEFAULTS["composer"])
    if args.narrator:
        set_cell(ws, row, "narrator", args.narrator)
    set_cell(ws, row, "copyright", args.copyright_text or DEFAULTS["copyright"])
    set_cell(ws, row, "year", int(args.year or DEFAULTS["year"]))

    # Metadata
    album = args.album or ALBUM_DEFAULTS.get(media_type, DEFAULTS["album"])
    set_cell(ws, row, "album", album)
    set_cell(ws, row, "genre", args.genre or DEFAULTS["genre"])
    if args.bpm:
        set_cell(ws, row, "bpm", int(args.bpm))
    if args.key:
        set_cell(ws, row, "key", args.key)

    # Source & Production
    if args.source_platform:
        set_cell(ws, row, "source_platform", args.source_platform)
    if args.source_file:
        set_cell(ws, row, "source_file", args.source_file)
    if args.mastered:
        set_cell(ws, row, "mastered", args.mastered)
    if args.mastering_tool:
        set_cell(ws, row, "mastering_tool", args.mastering_tool)

    # Tiers
    for tier in ["archive_wav", "distro_wav", "web_mp3", "album_art"]:
        set_cell(ws, row, tier, getattr(args, tier, None) or "—")
    set_cell(ws, row, "vbr_v2", args.vbr_v2 or "N")

    # Deployment
    if args.web_filename:
        set_cell(ws, row, "web_filename", args.web_filename)
    for dep in ["website_live", "r2_cdn", "distrokid", "streaming_live"]:
        set_cell(ws, row, dep, getattr(args, dep, None) or "—")
    if args.release_date:
        set_cell(ws, row, "release_date", args.release_date)

    # Pipeline
    set_cell(ws, row, "pipeline", args.pipeline or "Not Started")
    set_cell(ws, row, "id3_tagged", args.id3_tagged or "—")
    stamp_updated(ws, row)
    if args.notes:
        set_cell(ws, row, "notes", args.notes)

    # Apply styling to entire row
    for c in range(1, 40):
        apply_cell_style(ws, row, c)

    wb.save(catalog_path)
    print(f"✓ Added {cat_id}: \"{args.title}\" ({media_type}) at row {row}")
    print(f"  Catalog: {catalog_path}")


def cmd_update(args):
    """Update an existing entry by Catalog ID."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path)
    ws = wb["Media Catalog"]

    row = find_row_by_id(ws, args.catalog_id)
    if not row:
        print(f"Error: Catalog ID '{args.catalog_id}' not found.")
        sys.exit(1)

    _apply_updates(ws, row, args)
    stamp_updated(ws, row)

    wb.save(catalog_path)
    title = get_cell(ws, row, "title")
    print(f"✓ Updated {args.catalog_id}: \"{title}\" at row {row}")


def cmd_update_by_title(args):
    """Update an existing entry by title (+ optional style)."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path)
    ws = wb["Media Catalog"]

    row = find_row_by_title_style(ws, args.title, args.style)
    if not row:
        match_desc = f"'{args.title}'"
        if args.style:
            match_desc += f" / '{args.style}'"
        print(f"Error: No entry found matching {match_desc}.")
        sys.exit(1)

    _apply_updates(ws, row, args)
    stamp_updated(ws, row)

    wb.save(catalog_path)
    cat_id = get_cell(ws, row, "catalog_id")
    print(f"✓ Updated {cat_id}: \"{args.title}\" at row {row}")


def _apply_updates(ws, row: int, args):
    """Apply all non-None arg values to the given row."""
    # Map CLI arg names to column names
    field_map = {
        "isrc": "isrc", "media_type": "media_type", "prefix": "prefix",
        "chapter": "chapter", "track": "track", "version": "version",
        "title": "title", "style": "style", "duration": "duration",
        "description": "description", "artist": "artist", "composer": "composer",
        "narrator": "narrator", "copyright_text": "copyright", "year": "year",
        "album": "album", "genre": "genre", "bpm": "bpm", "key": "key",
        "source_platform": "source_platform", "source_file": "source_file",
        "mastered": "mastered", "mastering_tool": "mastering_tool",
        "archive_wav": "archive_wav", "distro_wav": "distro_wav",
        "web_mp3": "web_mp3", "vbr_v2": "vbr_v2", "album_art": "album_art",
        "web_filename": "web_filename", "website_live": "website_live",
        "r2_cdn": "r2_cdn", "distrokid": "distrokid",
        "release_date": "release_date", "streaming_live": "streaming_live",
        "pipeline": "pipeline", "id3_tagged": "id3_tagged", "notes": "notes",
    }
    updated = []
    for arg_name, col_name in field_map.items():
        val = getattr(args, arg_name, None)
        if val is not None:
            # Skip title/style for update-by-title (used as lookup keys)
            if hasattr(args, "catalog_id"):
                pass  # update command — apply everything
            elif arg_name in ("title",) and not hasattr(args, "new_title"):
                continue  # title was used as lookup, don't re-set
            # Type coercion
            if col_name in ("chapter", "track", "version", "year", "bpm"):
                val = int(val)
            set_cell(ws, row, col_name, val)
            apply_cell_style(ws, row, COL[col_name])
            updated.append(col_name)

    if updated:
        print(f"  Fields updated: {', '.join(updated)}")


def cmd_list(args):
    """List catalog entries with optional filters."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path, data_only=True)
    ws = wb["Media Catalog"]

    entries = []
    for row in range(get_data_start_row(), ws.max_row + 1):
        cat_id = get_cell(ws, row, "catalog_id")
        if not cat_id:
            continue

        # Apply filters
        if args.prefix:
            row_prefix = get_cell(ws, row, "prefix")
            if str(row_prefix or "").strip() != args.prefix.strip():
                continue
        if args.pipeline:
            row_pipeline = get_cell(ws, row, "pipeline")
            if args.pipeline.lower() not in str(row_pipeline or "").lower():
                continue
        if args.media_type:
            row_mt = get_cell(ws, row, "media_type")
            if str(row_mt or "").strip().lower() != args.media_type.strip().lower():
                continue

        entries.append({
            "id": cat_id,
            "type": get_cell(ws, row, "media_type") or "—",
            "prefix": get_cell(ws, row, "prefix") or "—",
            "ch": get_cell(ws, row, "chapter") or "",
            "title": get_cell(ws, row, "title") or "—",
            "style": get_cell(ws, row, "style") or "",
            "pipeline": get_cell(ws, row, "pipeline") or "—",
            "web": get_cell(ws, row, "website_live") or "—",
            "dk": get_cell(ws, row, "distrokid") or "—",
        })

    if not entries:
        print("No entries found matching filters.")
        return

    # Print table
    print(f"\n{'ID':<10} {'Type':<10} {'Pfx':<5} {'Ch':<4} {'Title':<34} "
          f"{'Style':<22} {'Pipeline':<20} {'Web':<5} {'DK':<5}")
    print("─" * 120)
    for e in entries:
        title = str(e['title'])[:32]
        style = str(e['style'])[:20]
        pipeline = str(e['pipeline'])[:18]
        print(f"{e['id']:<10} {str(e['type']):<10} {str(e['prefix']):<5} "
              f"{str(e['ch']):<4} {title:<34} {style:<22} "
              f"{pipeline:<20} {str(e['web']):<5} {str(e['dk']):<5}")

    print(f"\n{len(entries)} entries found.")


def cmd_next_id(args):
    """Show the next available Catalog ID."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path)
    ws = wb["Media Catalog"]
    print(next_catalog_id(ws))


def cmd_export(args):
    """Export the catalog to CSV."""
    catalog_path = find_catalog()
    wb = load_workbook(catalog_path, data_only=True)
    ws = wb["Media Catalog"]

    import csv
    output = args.output or "catalog_export.csv"
    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # Header row (row 2 has the column names)
        headers = [ws.cell(row=2, column=c).value for c in range(1, 40)]
        writer.writerow(headers)
        # Data rows
        count = 0
        for row in range(get_data_start_row(), ws.max_row + 1):
            cat_id = ws.cell(row=row, column=COL["catalog_id"]).value
            if not cat_id:
                continue
            row_data = [ws.cell(row=row, column=c).value for c in range(1, 40)]
            writer.writerow(row_data)
            count += 1

    print(f"✓ Exported {count} entries to {output}")


# ===========================================================================
# CLI ARGUMENT PARSER
# ===========================================================================
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="WoP Media Catalog Manager — add, update, and query entries.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", help="Command to run")

    # --- Shared field arguments ---
    def add_field_args(p, required_title=True):
        """Add common field arguments to a subparser."""
        # Identity
        p.add_argument("--isrc", help="ISRC code")
        p.add_argument("--media-type", dest="media_type",
                        help="Music, Narration, Podcast, Spoken Word, Ambient, Video, Other")
        p.add_argument("--prefix", help="Prefix code: 01-63, AN, SY, AM, SP, OV, NR, PO, SW, VD")
        p.add_argument("--chapter", help="Chapter number (1-63)")
        p.add_argument("--track", help="Track number within prefix group")
        p.add_argument("--version", help="Version number (default: 1)")
        # Content
        if required_title:
            p.add_argument("--title", required=True, help="Track/entry title")
        else:
            p.add_argument("--title", help="Track/entry title")
        p.add_argument("--style", help="Style descriptor (e.g., Sacred Americana, Gospel Soul)")
        p.add_argument("--duration", help="Duration in M:SS format")
        p.add_argument("--description", help="Brief description")
        # Credits
        p.add_argument("--artist", help=f"Artist (default: {DEFAULTS['artist']})")
        p.add_argument("--composer", help=f"Composer (default: {DEFAULTS['composer']})")
        p.add_argument("--narrator", help="Narrator / Host (for non-music)")
        p.add_argument("--copyright", dest="copyright_text",
                        help=f"Copyright (default: {DEFAULTS['copyright']})")
        p.add_argument("--year", help=f"Year (default: {DEFAULTS['year']})")
        # Metadata
        p.add_argument("--album", help="Album / Series name")
        p.add_argument("--genre", help=f"Genre (default: {DEFAULTS['genre']})")
        p.add_argument("--bpm", help="BPM")
        p.add_argument("--key", help="Musical key (e.g., G, Am)")
        # Source
        p.add_argument("--source-platform", dest="source_platform",
                        help="Suno AI, ElevenLabs, Kits AI, Live Recording, etc.")
        p.add_argument("--source-file", dest="source_file", help="Source filename")
        p.add_argument("--mastered", help="Y, N, Pending, N/A")
        p.add_argument("--mastering-tool", dest="mastering_tool",
                        help="Masterchannel, Manual (Audacity), etc.")
        # Tiers
        p.add_argument("--archive-wav", dest="archive_wav", help="✓, —, Pending")
        p.add_argument("--distro-wav", dest="distro_wav", help="✓, —, Pending")
        p.add_argument("--web-mp3", dest="web_mp3", help="✓, —, Pending")
        p.add_argument("--vbr-v2", dest="vbr_v2", help="Y or N")
        p.add_argument("--album-art", dest="album_art", help="✓, —, Pending")
        # Deployment
        p.add_argument("--web-filename", dest="web_filename", help="Filename in src/assets/audio/")
        p.add_argument("--website-live", dest="website_live", help="✓ or —")
        p.add_argument("--r2-cdn", dest="r2_cdn", help="✓, —, Pending")
        p.add_argument("--distrokid", help="✓, —, N/A")
        p.add_argument("--release-date", dest="release_date", help="YYYY-MM-DD")
        p.add_argument("--streaming-live", dest="streaming_live", help="✓, —, N/A")
        # Pipeline
        p.add_argument("--pipeline", help="Pipeline status")
        p.add_argument("--id3-tagged", dest="id3_tagged", help="✓ or —")
        p.add_argument("--notes", help="Notes")

    # ADD
    add_cmd = sub.add_parser("add", help="Add a new catalog entry")
    add_field_args(add_cmd, required_title=True)
    add_cmd.add_argument("--force", action="store_true",
                          help="Add even if duplicate title+style exists")

    # UPDATE (by Catalog ID)
    update_cmd = sub.add_parser("update", help="Update entry by Catalog ID")
    update_cmd.add_argument("catalog_id", help="Catalog ID (e.g., WOP-0001)")
    add_field_args(update_cmd, required_title=False)

    # UPDATE-BY-TITLE
    ubt_cmd = sub.add_parser("update-by-title",
                              help="Update entry by matching title + optional style")
    add_field_args(ubt_cmd, required_title=True)

    # LIST
    list_cmd = sub.add_parser("list", help="List catalog entries")
    list_cmd.add_argument("--prefix", help="Filter by prefix")
    list_cmd.add_argument("--pipeline", help="Filter by pipeline status (partial match)")
    list_cmd.add_argument("--media-type", dest="media_type", help="Filter by media type")

    # NEXT-ID
    sub.add_parser("next-id", help="Show next available Catalog ID")

    # EXPORT
    export_cmd = sub.add_parser("export", help="Export catalog to CSV")
    export_cmd.add_argument("output", nargs="?", default="catalog_export.csv",
                             help="Output CSV path")

    return parser


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(0)

    commands = {
        "add": cmd_add,
        "update": cmd_update,
        "update-by-title": cmd_update_by_title,
        "list": cmd_list,
        "next-id": cmd_next_id,
        "export": cmd_export,
    }

    cmd_func = commands.get(args.command)
    if cmd_func:
        cmd_func(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
