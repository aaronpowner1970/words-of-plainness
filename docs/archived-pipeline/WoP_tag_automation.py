#!/usr/bin/env python3
"""
WoP Publication Pipeline — ID3 Tag Automation
===============================================
Reads the WoP_ISRC_Catalog.xlsx spreadsheet and batch-applies ID3 tags
to MP3 and WAV files. Embeds album art. Updates the spreadsheet.

SETUP (one time):
    pip install mutagen openpyxl Pillow

USAGE:
    python wop_tag_automation.py                          # Interactive mode
    python wop_tag_automation.py --audio-dir ./audio      # Specify audio folder
    python wop_tag_automation.py --catalog ./catalog.xlsx  # Specify spreadsheet
    python wop_tag_automation.py --art ./cover.jpg         # Specify album art
    python wop_tag_automation.py --dry-run                 # Preview without writing

WHAT IT DOES:
    1. Reads track metadata from WoP_ISRC_Catalog.xlsx (Sheet: "Track Catalog")
    2. Finds matching audio files in the audio directory
    3. Embeds ID3v2.4 tags (MP3) or ID3 tags in WAV files
    4. Embeds album art (3000x3000 JPEG) if provided
    5. Updates the spreadsheet checkboxes for completed tracks
"""

import os
import sys
import argparse
import glob
from datetime import datetime

try:
    from mutagen.mp3 import MP3
    from mutagen.wave import WAVE
    from mutagen.id3 import (
        ID3, TIT2, TPE1, TPE2, TALB, TRCK, TDRC, TCON,
        COMM, TCOM, TCOP, TXXX, APIC, ID3NoHeaderError
    )
    MUTAGEN_OK = True
except ImportError:
    MUTAGEN_OK = False

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from PIL import Image
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False


# ── CONSTANTS ───────────────────────────────────────────────────────
COLUMN_MAP = {
    "A": "isrc",
    "B": "filename",
    "C": "title",
    "D": "prefix",
    "E": "chapter",
    "F": "version",
    "G": "style",
    "H": "artist",
    "I": "album",
    "J": "track_num",
    "K": "year",
    "L": "genre",
    "M": "composer",
    "N": "copyright",
    "O": "release_date",
    "P": "album_art",
    "Q": "archive_wav",
    "R": "distro_wav",
    "S": "web_mp3",
    "T": "website_live",
    "U": "distrokid",
    "V": "notes",
}

DEFAULTS = {
    "artist": "Words of Plainness",
    "album_artist": "Words of Plainness",
    "album": "Words of Plainness: Musical Testimonies",
    "year": "2026",
    "genre": "Christian / Sacred",
    "composer": "Aaron J Powner",
    "copyright": "© 2026 Aaron J Powner",
    "comment": "A Christ-Centered Ministry — words-of-plainness.vercel.app",
}


# ── SPREADSHEET READING ────────────────────────────────────────────
def read_catalog(catalog_path):
    """Read track metadata from the ISRC catalog spreadsheet."""
    wb = load_workbook(catalog_path, data_only=True)
    ws = wb["Track Catalog"]

    tracks = []
    for row in range(2, ws.max_row + 1):
        title = ws[f"C{row}"].value
        if not title:
            continue

        track = {"_row": row}
        for col, field in COLUMN_MAP.items():
            val = ws[f"{col}{row}"].value
            track[field] = str(val).strip() if val is not None else ""

        # Use defaults for empty fields
        for key, default in DEFAULTS.items():
            if key in track and not track[key]:
                track[key] = default

        tracks.append(track)

    wb.close()
    return tracks


def update_catalog(catalog_path, row, updates):
    """Update specific cells in the catalog spreadsheet."""
    wb = load_workbook(catalog_path)
    ws = wb["Track Catalog"]
    for col, value in updates.items():
        ws[f"{col}{row}"] = value
    wb.save(catalog_path)
    wb.close()


# ── FILE DISCOVERY ──────────────────────────────────────────────────
def find_audio_file(audio_dir, filename_base):
    """Find matching audio files (mp3 or wav) for a catalog entry."""
    matches = []
    if not filename_base:
        return matches

    for ext in ["*.mp3", "*.wav", "*.MP3", "*.WAV"]:
        for filepath in glob.glob(os.path.join(audio_dir, ext)):
            basename = os.path.splitext(os.path.basename(filepath))[0]
            # Match exactly or with close similarity
            if basename == filename_base or basename.replace(" ", "_") == filename_base:
                matches.append(filepath)

    # Also try partial matching (filename starts with the expected base)
    if not matches:
        for ext in ["*.mp3", "*.wav", "*.MP3", "*.WAV"]:
            for filepath in glob.glob(os.path.join(audio_dir, ext)):
                basename = os.path.splitext(os.path.basename(filepath))[0]
                if filename_base in basename or basename in filename_base:
                    matches.append(filepath)

    return matches


# ── ALBUM ART ───────────────────────────────────────────────────────
def load_album_art(art_path):
    """Load and validate album art. Returns JPEG bytes or None."""
    if not art_path or not os.path.exists(art_path):
        return None

    if PILLOW_OK:
        img = Image.open(art_path)
        w, h = img.size
        if w < 3000 or h < 3000:
            print(f"  ⚠  Album art is {w}x{h}px — recommended 3000x3000 minimum")
        if w != h:
            print(f"  ⚠  Album art is not square ({w}x{h}) — may display cropped")

    # Read raw bytes
    with open(art_path, "rb") as f:
        art_data = f.read()

    # Determine MIME type
    if art_path.lower().endswith(".png"):
        mime = "image/png"
    else:
        mime = "image/jpeg"

    return art_data, mime


# ── ID3 TAGGING ─────────────────────────────────────────────────────
def apply_tags_mp3(filepath, track, art_data=None, art_mime=None):
    """Apply ID3v2.4 tags to an MP3 file."""
    try:
        audio = MP3(filepath)
    except Exception:
        audio = MP3()

    # Ensure ID3 tags exist
    try:
        audio.add_tags()
    except Exception:
        pass

    tags = audio.tags

    # Core tags
    tags.add(TIT2(encoding=3, text=track.get("title", "")))
    tags.add(TPE1(encoding=3, text=track.get("artist", DEFAULTS["artist"])))
    tags.add(TPE2(encoding=3, text=DEFAULTS["album_artist"]))
    tags.add(TALB(encoding=3, text=track.get("album", DEFAULTS["album"])))
    tags.add(TCON(encoding=3, text=track.get("genre", DEFAULTS["genre"])))
    tags.add(TCOM(encoding=3, text=track.get("composer", DEFAULTS["composer"])))
    tags.add(TCOP(encoding=3, text=track.get("copyright", DEFAULTS["copyright"])))

    if track.get("track_num"):
        tags.add(TRCK(encoding=3, text=str(track["track_num"])))
    if track.get("year"):
        tags.add(TDRC(encoding=3, text=str(track["year"])))

    # Comment
    comment = track.get("comment", DEFAULTS["comment"])
    tags.add(COMM(encoding=3, lang="eng", desc="", text=comment))

    # ISRC as custom tag
    if track.get("isrc"):
        tags.add(TXXX(encoding=3, desc="ISRC", text=track["isrc"]))

    # Album art
    if art_data:
        tags.add(APIC(
            encoding=3,
            mime=art_mime or "image/jpeg",
            type=3,  # Cover (front)
            desc="Cover",
            data=art_data,
        ))

    audio.save()
    return True


def apply_tags_wav(filepath, track, art_data=None, art_mime=None):
    """Apply ID3 tags to a WAV file."""
    audio = WAVE(filepath)

    try:
        audio.add_tags()
    except Exception:
        pass

    tags = audio.tags

    tags.add(TIT2(encoding=3, text=track.get("title", "")))
    tags.add(TPE1(encoding=3, text=track.get("artist", DEFAULTS["artist"])))
    tags.add(TPE2(encoding=3, text=DEFAULTS["album_artist"]))
    tags.add(TALB(encoding=3, text=track.get("album", DEFAULTS["album"])))
    tags.add(TCON(encoding=3, text=track.get("genre", DEFAULTS["genre"])))
    tags.add(TCOM(encoding=3, text=track.get("composer", DEFAULTS["composer"])))
    tags.add(TCOP(encoding=3, text=track.get("copyright", DEFAULTS["copyright"])))

    if track.get("track_num"):
        tags.add(TRCK(encoding=3, text=str(track["track_num"])))
    if track.get("year"):
        tags.add(TDRC(encoding=3, text=str(track["year"])))

    tags.add(COMM(encoding=3, lang="eng", desc="", text=track.get("comment", DEFAULTS["comment"])))

    if track.get("isrc"):
        tags.add(TXXX(encoding=3, desc="ISRC", text=track["isrc"]))

    if art_data:
        tags.add(APIC(
            encoding=3,
            mime=art_mime or "image/jpeg",
            type=3,
            desc="Cover",
            data=art_data,
        ))

    audio.save()
    return True


# ── MAIN ────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="WoP Publication Pipeline — ID3 Tag Automation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python wop_tag_automation.py --audio-dir ./WoP-Audio/03-Web-MP3s
  python wop_tag_automation.py --art ./cover_3000x3000.jpg --dry-run
  python wop_tag_automation.py --catalog ./WoP_ISRC_Catalog.xlsx --audio-dir ./audio
        """,
    )
    parser.add_argument("--catalog", default="WoP_ISRC_Catalog.xlsx",
                        help="Path to ISRC catalog spreadsheet (default: WoP_ISRC_Catalog.xlsx)")
    parser.add_argument("--audio-dir", default=".",
                        help="Directory containing audio files (default: current directory)")
    parser.add_argument("--art", default=None,
                        help="Path to album art image (JPEG/PNG, 3000x3000 recommended)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview what would be done without writing files")
    parser.add_argument("--update-spreadsheet", action="store_true",
                        help="Update spreadsheet checkboxes after tagging")

    args = parser.parse_args()

    # ── Dependency check ──
    if not MUTAGEN_OK:
        print("ERROR: mutagen not installed. Run: pip install mutagen")
        sys.exit(1)
    if not OPENPYXL_OK:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    # ── Load catalog ──
    if not os.path.exists(args.catalog):
        print(f"ERROR: Catalog not found: {args.catalog}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  WoP Publication Pipeline — ID3 Tag Automation")
    print(f"{'='*60}")
    print(f"  Catalog:   {args.catalog}")
    print(f"  Audio dir: {os.path.abspath(args.audio_dir)}")
    print(f"  Album art: {args.art or 'None'}")
    print(f"  Dry run:   {args.dry_run}")
    print(f"{'='*60}\n")

    tracks = read_catalog(args.catalog)
    print(f"Found {len(tracks)} tracks in catalog.\n")

    if not tracks:
        print("No tracks found. Add entries to the Track Catalog sheet.")
        sys.exit(0)

    # ── Load album art ──
    art_data, art_mime = None, None
    if args.art:
        result = load_album_art(args.art)
        if result:
            art_data, art_mime = result
            print(f"✓ Album art loaded: {args.art}\n")
        else:
            print(f"⚠ Could not load album art: {args.art}\n")

    # ── Process tracks ──
    tagged = 0
    skipped = 0
    not_found = 0

    for track in tracks:
        filename_base = track.get("filename", "").strip()
        title = track.get("title", "Unknown")

        if not filename_base:
            print(f"  ⏭  [{title}] — No filename in catalog, skipping")
            skipped += 1
            continue

        # Find audio files
        matches = find_audio_file(args.audio_dir, filename_base)

        if not matches:
            print(f"  ✗  [{title}] — File not found: {filename_base}.*")
            not_found += 1
            continue

        for filepath in matches:
            ext = os.path.splitext(filepath)[1].lower()
            basename = os.path.basename(filepath)

            if args.dry_run:
                print(f"  🔍 [{title}] — Would tag: {basename}")
                print(f"       Artist: {track.get('artist', DEFAULTS['artist'])}")
                print(f"       Album:  {track.get('album', DEFAULTS['album'])}")
                print(f"       Genre:  {track.get('genre', DEFAULTS['genre'])}")
                print(f"       Art:    {'Yes' if art_data else 'No'}")
                tagged += 1
                continue

            try:
                if ext == ".mp3":
                    apply_tags_mp3(filepath, track, art_data, art_mime)
                elif ext == ".wav":
                    apply_tags_wav(filepath, track, art_data, art_mime)
                else:
                    print(f"  ⏭  [{title}] — Unsupported format: {ext}")
                    skipped += 1
                    continue

                print(f"  ✓  [{title}] — Tagged: {basename}")
                tagged += 1

                # Update spreadsheet
                if args.update_spreadsheet:
                    row = track["_row"]
                    if ext == ".mp3":
                        update_catalog(args.catalog, row, {"S": "✓"})
                    elif ext == ".wav":
                        # Check which WAV tier based on directory
                        if "archive" in args.audio_dir.lower() or "01-" in args.audio_dir:
                            update_catalog(args.catalog, row, {"Q": "✓"})
                        elif "distro" in args.audio_dir.lower() or "02-" in args.audio_dir:
                            update_catalog(args.catalog, row, {"R": "✓"})
                    if art_data:
                        update_catalog(args.catalog, row, {"P": "✓ Embedded"})

            except Exception as e:
                print(f"  ✗  [{title}] — Error: {e}")

    # ── Summary ──
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  {'Tagged' if not args.dry_run else 'Would tag'}: {tagged}")
    print(f"  Skipped:    {skipped}")
    print(f"  Not found:  {not_found}")
    if args.dry_run:
        print(f"\n  This was a dry run. Add --no-dry-run to apply tags.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
